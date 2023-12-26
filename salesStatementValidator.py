

#PARAMS
#{ 
#   "ARGUMENTS": [ 
#       { "KEY": "-spec",    "TYPE": "openFile",     "FORMAT": "", "DESCR": "Спецификация",                  "DEFAULT": "Спецификация******.xlsx", "_RT_USAGE": "RES_FILE" },
#       { "KEY": "-odinAss", "TYPE": "openDir",      "FORMAT": "", "DESCR": "Папка с выгрузкой из 1C",       "DEFAULT": "C:\\SomeFolder\\...." },
#       { "KEY": "-verbose", "TYPE": "list(Да|Нет)", "FORMAT": "", "DESCR": "Выводить сообщения об ошибках", "DEFAULT": "Да" }
#       ],
#   "REQUIREMENTS": "pandas\nopenpyxl\nxlwings"
#}
#ENDPARAMS


#region  Library

import pandas as pd
import os
import openpyxl

class excelTableParser(pd.Series):
    def __init__(self, projectFileName:str, vectorsDescription: dict) -> None:

        #print(f'excelTableParser v0.1.0 2023/12/06')

        super().__init__()

        self.vType = vectorsDescription

        self.excelFileName = projectFileName
        self.xls = pd.ExcelFile(projectFileName)

    def closeXlsx(self):
        try:
            self.xls.close()
        except:
            pass

    def __del__(self):
        self.closeXlsx()

    def getVectors(self, vectors: list[str] = None):
        for vector in vectors:
            if vector in self.vType.keys():

                if vector not in self.keys():

                    sVECTOR = self.xls.parse(sheet_name = self.vType[vector]["Sheet"], index_col=None, skiprows=self.vType[vector]["SkipRows"], na_values='') 
                    sVECTOR_L1 = sVECTOR.dropna(subset=self.vType[vector]["L1_Col"], ).fillna('')
                    sVECTOR_L1 = sVECTOR_L1.rename(columns=self.vType[vector]["L1"])

                    if self.vType[vector]["L2"] is not None:
                        sVECTOR_L1["vElements"] = None
                        sVECTOR.drop(columns = set.difference(set(sVECTOR.keys()), set(self.vType[vector]["L2"].keys())), inplace = True)

                    if self.vType[vector]["L2"] is not None:
                        for i, idx in enumerate(sVECTOR_L1.index):
                            if i < len(sVECTOR_L1.index)-1:
                                sVECTOR_L1.at[idx,"vElements"] = sVECTOR.iloc[idx:sVECTOR_L1.index[i+1]].rename(columns=self.vType[vector]["L2"]).dropna(subset=self.vType[vector]["L2_Col"])
                            else:
                                sVECTOR_L1.at[idx,"vElements"] = sVECTOR.iloc[idx:].rename(columns=self.vType[vector]["L2"]).dropna(subset=self.vType[vector]["L2_Col"])

                    sVECTOR_L1.drop(columns = set.difference(set(sVECTOR.keys()), set(self.vType[vector]["L1"].keys())), inplace = True)

                    self[vector] = sVECTOR_L1.copy()
            else:
                print(f"WARNING!!! Vector {vector} undefined.")

        return self

    # Метод для сохранения изменённых полей векторов
    # update(writeVectors: dict)
    @staticmethod
    def getSheetNumber(wb: openpyxl.Workbook, listName: str):
        res = [i for i,x in enumerate(wb.sheetnames) if x == listName]
        if len(res) == 0:
            return None
        else:
            return res[0]

    @staticmethod
    def getColByName(ws: openpyxl.worksheet, Name: str, row: int = 1) -> int:
        for i in range(300):
            if ws.cell(row = row, column=i+1).value == Name:
                return i+1
        return None

    @staticmethod
    def eq(val1, val2):

        if val1 == val2:
            return True

        try:
            if int(val1) == int(val2):
                return True
        except:
            pass

        try:
            if float(val1) == float(val2):
                return True
        except:
            pass

        return False

    @staticmethod
    def getRowById(ws: openpyxl.worksheet, ID_Col: str, ID_Value: str, idxRow: int = 1) -> int:
        colNum = excelTableParser.getColByName(ws, ID_Col, idxRow)
        emptyCtr = 0
        for row in range(25000):
            val = ws.cell(row = row+1, column = colNum).value
            if val in ("", None):
                emptyCtr += 1
                if emptyCtr >= 50:
                    return None
            else:
                emptyCtr = 0

            if excelTableParser.eq(val, ID_Value):
                return row+1
        return None

    def update(self, writeVectors: dict):

        wb = openpyxl.load_workbook(self.excelFileName)

        for vector in writeVectors:
            ws = wb.worksheets[self.getSheetNumber(wb, self.vType[vector]["Sheet"])]

            for field in writeVectors[vector]:
                print(f'Write data for {vector}:{field}')
                col = self.getColByName(ws, field, self.vType[vector]["SkipRows"]+1)
                for row in self[vector].iloc:
                    xlsRow = self.getRowById(ws, self.vType[vector]["L1_Col"], row[self.vType[vector]["L1_Col"]], self.vType[vector]["SkipRows"]+1)
                    ws.cell(row=xlsRow, column = col, value = row[field])

        wb.close()
        wb.save(self.excelFileName)

#endregion

lastSource = None
def printError(source, error):
    global lastSource, arguments
    
    if arguments.verbose != 'Да':
        return
    
    if lastSource is None or source != lastSource:
        print(f'\r\n"{source}">>>')
    print(f'    {error}')

    lastSource = source

symbolsMatch = { 
'A': 'А', 'B': 'В', 'C': 'С', 'E': 'Е', 'H': 'Н', 'K': 'К', 'M': 'М', 'O': 'О', 'P': 'Р', 'T': 'Т', 'X': 'Х', #upper
'a': 'а', 'c': 'с', 'e': 'е', 'k': 'к', 'm': 'м', 'n': 'п', 'o': 'о', 'p': 'р', 'r': 'г', 'x': 'х',           #lower
'-': '', '_': '', '–': '', ' ': ''                                                                            #special
    }

def replaceByDict(text: str, rule: dict) -> str:
    res = str(text)
    for key in rule:
        res = res.replace(key, rule[key])
    return trimPrefixZeros(res.strip())

def trimPrefixZeros(text: str):
    if (len(text)) == 0:
        return text
    
    while(text[0] == '0'):
        text = text[1:]
        
    return text

def compareStr(str1: str, str2: str, rule: dict) -> bool: # bool, bool: Equal (Like), Equal and not mistype (Equal)
    if str1 == str2:
        return True, True
    elif replaceByDict(str1, rule) == replaceByDict(str2, rule):
        return True, False

    return False, False

def isStrLike(str1: str, str2: str, rule) -> bool:
    s1 = replaceByDict(str1, rule)
    s2 = replaceByDict(str2, rule)

    return s1 in s2 or s2 in s1

def writeComment(sheet, index: int, comment: str):
    sheet.iat[index, 4] = comment

try:

    typicalSheet_Spec = {
        "Sheet": "", # Subject to change
        "L1_Col": "Краткое описание",
        "L1": {
           '№' : '№', 'Краткое описание' : 'Краткое описание', 'Заказной номер' : 'Заказной номер',
           'Количество' : 'Количество', 'Комментарий': 'Комментарий'
           },
        "L2_Col": None,#"Краткое описание",
        "L2": None,#{'Краткое описание' : 'Краткое описание', 'Заказной номер' : 'Заказной номер', 'Количество' : 'Количество'},
        "SkipRows": 1
        }

    typicalSheet_OdinAss = {
        "Sheet": "TDSheet",
        "L1_Col": "Артикул",
        "L1": {
           '№' : '№', 'Артикул' : 'Артикул', 'Товар' : 'Товар', 'Кол-во' : 'Кол-во', 'Ед.': 'Ед.',
           'Ед. в штуках': 'Ед. в штуках'
           },
        "L2_Col": None,
        "L2": None,
        "SkipRows": 1 # Subject to change
        }

    from ScriptRuntimeData import arguments

    ## Копируем спецификацию
    import shutil
    import ntpath
    rawSpecName = ntpath.basename(arguments.spec);
    print(f'Копируем "{rawSpecName}" => "{"Обработка_"+rawSpecName}"')
    newFileName = os.path.join(os.path.dirname(arguments.spec), "Обработка_"+rawSpecName)
    shutil.copyfile(arguments.spec, newFileName)

    # Добавляем нужные столбцы
    wb = openpyxl.load_workbook(newFileName)
    for i in range(len(wb.sheetnames)):
        ws = wb.worksheets[i]
        ws.cell(row = 2, column=5, value="Комментарий")
    wb.save(newFileName)
    wb.close()

    # Читаем спецификацию
    print("Читаем спецификацию...")

    sheetNames_Spec = pd.ExcelFile(newFileName).sheet_names
    dct = {}
    for sheet in sheetNames_Spec:
        dc = typicalSheet_Spec.copy()
        dc["Sheet"] = sheet
        dct[sheet] = dc
    
    xlsDf_Spec = excelTableParser(newFileName, dct)
    xlsDf_Spec.getVectors(sheetNames_Spec)

    import xlwings as xw
    def findOdinAssSheetFirstRow(filename: str) -> int:

        try:
            ws = openpyxl.load_workbook(filename).worksheets[0]
        except:
            wb = xw.Book(filename); wb.save(); wb.close()
            ws = openpyxl.load_workbook(filename).worksheets[0]

        for row in range(40):
            for col in range(50):
                if ws.cell(row = row+1, column=col+1).value == "№":
                    return row
        return None
    
    # Читаем все файлы выгрузки ОдинЭс
    print("Читаем все файлы выгрузки 1С...")

    xlsDfs_OdinAss = {}
    for file in [x for x in os.listdir(arguments.odinAss) if x.endswith(".xlsx")]:
        
        fullName = os.path.join(arguments.odinAss, file)

        fRow = findOdinAssSheetFirstRow(fullName)

        try:

            print(f'    "{file}">>>')

            if fRow is not None:
                dc = {}
                dc["TDSheet"] = typicalSheet_OdinAss.copy()
                dc["TDSheet"]["SkipRows"] = fRow
                xlsDfs_OdinAss[file] = excelTableParser(fullName, dc)
                xlsDfs_OdinAss[file].getVectors(["TDSheet"])
            else:
                print(f"Не найдена корректная таблица в файле: \"{file}\"")

        except:

            print(f'        "Некорректный формат файла')

    # Сравниваем все листы Спецификации с соответствующими Файлами 1С (Лист: "{name}" <-> Файл "{name}.xlsx")
    mess = 'Сравниваем все листы Спецификации с соответствующими Файлами 1С (Лист: "{name}" <-> Файл "{name}.xlsx")'
    print("\n" + "#" * len(mess))
    print('Сравниваем все листы Спецификации с соответствующими Файлами 1С (Лист: "{name}" <-> Файл "{name}.xlsx")')
    print("#" * len(mess))
    
    missingInSpec = {}
    dfFullSpec = pd.DataFrame(columns=["Производитель", "Заказной номер", "OrderNum", "Требуется", "Описание", "В наличии", "Разница"])
    filesDict = [x for x in os.listdir(arguments.odinAss) if x.endswith(".xlsx")]
    for sheet in xlsDf_Spec.keys():

        if arguments.verbose != "Да":
            print(sheet)



        mtc = [x for x in xlsDfs_OdinAss if x == sheet+".xlsx"]
        if len(mtc) != 1:
            from openpyxl.styles import Font
            print(f"Не найден файл, соответствующий листу {sheet}!")
            wb = openpyxl.load_workbook(newFileName)
            ws = wb.worksheets[excelTableParser.getSheetNumber(wb, sheet)]
            ws.cell(row=1, column = 1, value = "Не найдены данные о закупках")
            ws["A1"].font = Font(color="FF0000", bold=True, size=20)
            wb.save(newFileName)
            wb.close()

            df_1C = pd.DataFrame({'№': [], 'Артикул': [], 'Товар': [], 'Кол-во': [], 'Ед.': [], 'Ед. в штуках': []})
        else:
            df_1C = xlsDfs_OdinAss[mtc[0]]["TDSheet"]

        lastVendor = "NA"

        df_1C_copy = df_1C.copy()
        for i, item in enumerate(xlsDf_Spec[sheet].iloc): # Спецификация

            if item["№"] is not None and len(item["№"].strip()) > 0:
                lastVendor = item["№"].strip()

            matchIn1C = None
            likeIn1C = None
            for item_1C in df_1C.iloc:

                strLike, strEq = compareStr(str(item["Заказной номер"]), str(item_1C["Артикул"]), symbolsMatch)
                if strLike:
                    matchIn1C = {'Артикул': item_1C["Артикул"], "Кол-во": item_1C["Кол-во"]}

                    df_1C_copy = df_1C_copy[df_1C_copy['Артикул'] != item_1C["Артикул"]] # Удаляем из вектора 1С найденный элемент

                    if not strEq:
                        printError(sheet, f'"{lastVendor}"/"{item["Заказной номер"]}">>> Неточное соответствие заказного номера')
                        writeComment(xlsDf_Spec[sheet], i, f'Неточное соответствие заказного номера, в 1С: "{str(item_1C["Артикул"])}"')

                    if (item["Количество"] != item_1C["Кол-во"]):
                        printError(sheet, f'"{lastVendor}"/"{item["Заказной номер"]}">>> Не соответствует количество (есть: {item_1C["Кол-во"]}, надо: {item["Количество"]})')
                        writeComment(xlsDf_Spec[sheet], i, f'Не соответствует количество (есть: {item_1C["Кол-во"]}, надо: {item["Количество"]})')

                else:

                    if isStrLike(str(item["Заказной номер"]), str(item_1C["Артикул"]), symbolsMatch):
                        likeIn1C = str(item_1C["Артикул"])

                if matchIn1C is not None:
                    continue # Если  нашли соответствие, обрабатываем следующие записи из спецификации

            if matchIn1C is None:

                ps = ""
                if likeIn1C is not None:
                    ps = f' Найден похожий артикул: "{likeIn1C}"'

                printError(sheet, f'"{lastVendor}"/"{item["Заказной номер"]}">>> Не найдено соответствия в компл. БД')
                writeComment(xlsDf_Spec[sheet], i, "Не найдено соответствия в компл. БД." + ps)
                
                
            # Заполнение сводного отчёта
            orderNumUnified = replaceByDict(item["Заказной номер"], symbolsMatch)
            newRecord = pd.DataFrame(
                {
                   "Производитель": [lastVendor], "Заказной номер": [item["Заказной номер"]], "Описание": [item["Краткое описание"]],
                   "OrderNum": [orderNumUnified], "Требуется": [0.0], "В наличии": [0.0], "Разница": [""]
                   })
                   
            if len(dfFullSpec.loc[dfFullSpec["OrderNum"] == orderNumUnified]) == 0:
                    if len(dfFullSpec) == 0:
                        dfFullSpec = newRecord.copy()
                    else:
                        dfFullSpec = pd.concat([dfFullSpec, newRecord])
            
            dfFullSpec.loc[dfFullSpec["OrderNum"] == orderNumUnified, "Требуется"] += item["Количество"]
            if matchIn1C is not None:
                dfFullSpec.loc[dfFullSpec["OrderNum"] == orderNumUnified, "В наличии"] += matchIn1C["Кол-во"]
                

        # Печатаем элементы из выгрузки 1С, оставшиеся ненайденными
        if len(df_1C_copy) > 0:
            if arguments.verbose == "Да":
                print("")
            missingInSpec[sheet] = []
            for item in df_1C_copy.iloc:
                printError(f'В файле "{sheet+".xlsx"}" найдены записи, соответствия которым в спецификации не обнаружено:', "(" + str(item['Артикул']) + ') ' + str(item['Товар']))
                missingInSpec[sheet].append((str(item['Артикул']), str(item['Товар'])))

    mess = f'Записываем комментарии об ошибках в файл "{newFileName}"'
    print("\n" + "#" * len(mess))
    print(mess)
    print("#" * len(mess))

    wv = {}
    for sheet in xlsDf_Spec.keys():
        wv[sheet] = ["Комментарий"]

    xlsDf_Spec.update(wv)

    if len(missingInSpec) > 0:

        print('\nПеречисляем отсутствующие в спецификации данные на листах:')
        wb = openpyxl.load_workbook(newFileName)

        for sheet in missingInSpec:
            startRow = len(xlsDf_Spec[sheet]) + 5

            print("    " + sheet)
            ws = wb.worksheets[excelTableParser.getSheetNumber(wb, sheet)]
            ws.cell(row = startRow, column=1, value="Кроме того найдено в 1С:")
            startRow += 1

            for i,item in enumerate(missingInSpec[sheet]):
                ws.cell(row = startRow+i, column=1, value=item[0])
                ws.cell(row = startRow+i, column=2, value=item[1])

        wb.save(newFileName)
        wb.close()
        
    print('\nСоздаём сводный отчёт...')

    dfFullSpec['Разница'] = dfFullSpec["В наличии"] - dfFullSpec["Требуется"]
    dfFullSpec.loc[dfFullSpec['Разница'] == 0, 'Разница'] = None
    dfFullSpec.sort_values(["Производитель", "Заказной номер"], inplace=True)
    dfFullSpec.reset_index(inplace=True)
    dfFullSpec.drop(columns=["OrderNum","index"], inplace=True)

    dfFullSpec.to_excel( os.path.join( ntpath.dirname(newFileName), "Сводный отчёт.xlsx" ), sheet_name= "Сводный" )

    print("\nГотово!")
    print("RETURN_SUCCESS")

except Exception as ex:

    print("RETURN_ERROR")
    print(ex)
    
